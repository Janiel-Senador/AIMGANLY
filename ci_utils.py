import re
from copy import deepcopy
from datetime import datetime
from html import unescape
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook


TEMPLATE_DEFINITIONS: Dict[str, Dict[str, Any]] = {
    "bpi": {
        "label": "BPI PDRN",
        "path": Path(r"C:\Users\Jan\Downloads\PDRN-MAPPING-1_efa98837-83fc-4d76-987a-3101659aafce.xls"),
        "sheet_name": "PDRN",
    },
    "maybank": {
        "label": "Maybank PDRN",
        "path": Path(__file__).resolve().parent / "maybank" / "pdrn" / "MAYBANK PDRN.xlsx",
        "sheet_name": "Sheet1",
    },
}

DEFAULT_TEMPLATE_KEY = "bpi"
DEFAULT_TEMPLATE_PATH = TEMPLATE_DEFINITIONS[DEFAULT_TEMPLATE_KEY]["path"]


CI_FORM_DEFAULTS: Dict[str, Any] = {
    "reference_number": "",
    "bpi_request_number": "",
    "los_request_number": "",
    "account_name": "",
    "subject_last_name": "",
    "subject_first_name": "",
    "subject_middle_name": "",
    "subject_maiden_middle_name": "",
    "subject_nickname": "",
    "subject_bday_age": "",
    "subject_birthplace": "",
    "subject_civil_status": "",
    "subject_years_married_or_separated": "",
    "subject_gender": "",
    "subject_nationality": "",
    "subject_education": "",
    "purpose_of_loan": "",
    "subject_employment": "",
    "subject_business": "",
    "subject_employer_name": "",
    "subject_employer_address": "",
    "subject_position": "",
    "subject_tenure": "",
    "subject_salary": "",
    "subject_employment_status": "",
    "subject_ofw_status": "",
    "subject_departure_date": "",
    "subject_arrival_date": "",
    "subject_business_name": "",
    "subject_business_address": "",
    "subject_business_years_of_operation": "",
    "subject_monthly_gross_income": "",
    "subject_monthly_net_income": "",
    "subject_monthly_remittance_amount": "",
    "subject_remittance_center": "",
    "subject_remittance_branch": "",
    "subject_sender_name": "",
    "spouse_last_name": "",
    "spouse_first_name": "",
    "spouse_middle_name": "",
    "spouse_maiden_middle_name": "",
    "spouse_nickname": "",
    "spouse_bday_age": "",
    "spouse_birthplace": "",
    "spouse_civil_status": "",
    "spouse_gender": "",
    "spouse_nationality": "",
    "spouse_education": "",
    "spouse_employment": "",
    "spouse_business": "",
    "spouse_employer_name": "",
    "spouse_employer_address": "",
    "spouse_position": "",
    "spouse_tenure": "",
    "spouse_salary": "",
    "spouse_employment_status": "",
    "spouse_business_name": "",
    "spouse_business_address": "",
    "spouse_business_years_of_operation": "",
    "spouse_monthly_gross_income": "",
    "spouse_monthly_net_income": "",
    "spouse_monthly_remittance_amount": "",
    "spouse_remittance_center": "",
    "spouse_remittance_branch": "",
    "spouse_sender_name": "",
    "spouse_relationship_to_sender": "",
    "dependents": "",
    "given_address": "",
    "verified_address": "",
    "complete_address": "",
    "present_address": "",
    "previous_address": "",
    "province": "",
    "region": "",
    "contact_number": "",
    "length_of_stay": "",
    "ownership": "",
    "landlord_name": "",
    "rental_fee": "",
    "vehicles": "",
    "vehicle_unit_year_model": "",
    "vehicle_owned_or_mortgage": "",
    "vehicle_seen_status": "",
    "vehicle_condition": "",
    "vehicle_monthly_amortization": "",
    "vehicle_years_to_pay": "",
    "vehicle_remaining_years_to_pay": "",
    "type_of_residence": "",
    "no_of_storey": "",
    "classification": "",
    "area": "",
    "house_condition": "",
    "made_of": "",
    "appearance": "",
    "accessibility": "",
    "neighborhood": "",
    "house_color": "",
    "gate_color": "",
    "fence_color": "",
    "interior": "",
    "exterior": "",
    "no_of_bedroom": "",
    "no_of_toilet_bath": "",
    "parking": "",
    "lot_area": "",
    "floor_area": "",
    "landmark": "",
    "corner": "",
    "security_guard_hoa": "",
    "time_of_visit": "",
    "nearest_bdo_branch": "",
    "bdo_distance": "",
    "utility_bills": "",
    "coordinates": "",
    "ownership_verification": "",
    "basis_of_ownership_verification": "",
    "living_condition": "",
    "neighborhood_classification": "",
    "area_definition": "",
    "adverse_finding": "",
    "ci_remarks_observation": "",
    "barangay_informant_name": "",
    "barangay_informant_position": "",
    "barangay_informant_age": "",
    "barangay_informant_contact_number": "",
    "barangay_informant_address": "",
    "barangay_subject_known": "",
    "barangay_subject_voter": "",
    "barangay_subject_resident": "",
    "barangay_has_bad_record": "",
    "barangay_same_address": "",
    "barangay_record_based_on": "",
    "barangay_selfie_picture": "",
    "barangay_logbook_picture": "",
    "barangay_voter_record_picture": "",
    "barangay_remarks": "",
    "neighbor_1_name": "",
    "neighbor_1_age": "",
    "neighbor_1_relationship": "",
    "neighbor_1_years_in_area": "",
    "neighbor_1_address": "",
    "neighbor_1_known": "",
    "neighbor_1_selfie_picture": "",
    "neighbor_1_remarks": "",
    "neighbor_2_name": "",
    "neighbor_2_age": "",
    "neighbor_2_relationship": "",
    "neighbor_2_years_in_area": "",
    "neighbor_2_address": "",
    "neighbor_2_known": "",
    "neighbor_2_selfie_picture": "",
    "neighbor_2_remarks": "",
    "neighbor_3_name": "",
    "neighbor_3_age": "",
    "neighbor_3_relationship": "",
    "neighbor_3_years_in_area": "",
    "neighbor_3_address": "",
    "neighbor_3_known": "",
    "neighbor_3_selfie_picture": "",
    "neighbor_3_remarks": "",
    "main_informant_name": "",
    "main_informant_address": "",
    "main_informant_age": "",
    "main_informant_relationship": "",
    "main_informant_selfie_picture": "",
    "main_informant_remarks": "",
    "remarks": "",
    "informants": "",
    "main_informant": "",
    "field_investigator": "",
    "requesting_officer": "",
    "date_time_inspected": "",
    "submitted_date": "",
    "submitted_time": "",
    "outcome": "",
    "raw_notes": "",
}


DIRECT_FIELD_MAP = {
    "ACCOUNT NAME": "account_name",
    "BDAY AGE": "subject_bday_age",
    "BIRTHPLACE": "subject_birthplace",
    "CIVIL STATUS": "subject_civil_status",
    "NATIONALITY": "subject_nationality",
    "EDUCATIONAL ATTAINMENT": "subject_education",
    "EMPLOYMENT": "subject_employment",
    "BUSINESS": "subject_business",
    "PURPOSE OF LOAN": "purpose_of_loan",
    "GIVEN ADDRESS": "given_address",
    "VERIFIED ADDRESS": "verified_address",
    "COMPLETE ADDRESS": "complete_address",
    "PRESENT ADDRESS": "present_address",
    "LENGTH OF STAY": "length_of_stay",
    "PREVIOUS": "previous_address",
    "PREVIOUS ADDRESS": "previous_address",
    "PROVINCE": "province",
    "CONTACT NUMBER TELEPHONE NUMBER": "contact_number",
    "CONTACT NUMBER": "contact_number",
    "OWNERSHIP": "ownership",
    "NAME OF LANDLORD": "landlord_name",
    "RENTAL FEE": "rental_fee",
    "VEHICLES": "vehicles",
    "UNIT YEAR MODEL": "vehicle_unit_year_model",
    "OWNED OR MORTGAGE": "vehicle_owned_or_mortgage",
    "SEEN NOT SEEN": "vehicle_seen_status",
    "CONDITION": "vehicle_condition",
    "MONTHLY AMORTIZATION IF MORTGAGED": "vehicle_monthly_amortization",
    "YEARS TO PAY": "vehicle_years_to_pay",
    "REMAINING YEARS TO PAY": "vehicle_remaining_years_to_pay",
    "TYPE OF RESIDENCE": "type_of_residence",
    "NO OF STOREY": "no_of_storey",
    "CLASSIFICATION": "classification",
    "AREA": "area",
    "HOUSE CONDITION": "house_condition",
    "MADE": "made_of",
    "APPEARANCE": "appearance",
    "ACCESSIBILITY": "accessibility",
    "NEIGHBORHOOD": "neighborhood",
    "HOUSE COLOR": "house_color",
    "GATE COLOR": "gate_color",
    "FENCE COLOR": "fence_color",
    "INTERIOR": "interior",
    "EXTERIOR": "exterior",
    "NO OF BEDROOM": "no_of_bedroom",
    "NO OF TOILET BATH": "no_of_toilet_bath",
    "LIVING CONDITION": "living_condition",
    "PARKING": "parking",
    "PARKING GARAGE": "parking",
    "LOT AREA": "lot_area",
    "FLOOR AREA": "floor_area",
    "LANDMARK AND METERS AWAY": "landmark",
    "LANDMARK": "landmark",
    "NEAREST CORNER": "corner",
    "CORNER": "corner",
    "IF WITH SECURITY GUARD AND HOA OFFICE": "security_guard_hoa",
    "TIME OF VISIT": "time_of_visit",
    "NEAREST BDO BRANCH": "nearest_bdo_branch",
    "KILOMETERS METERS AWAY": "bdo_distance",
    "CI REMARKS AND OBSERVATION": "ci_remarks_observation",
}


PERSON_FIELD_MAP = {
    "LAST NAME": "last_name",
    "FIRST NAME": "first_name",
    "MIDDLE NAME": "middle_name",
    "MAIDEN MIDDLE NAME": "maiden_middle_name",
    "NICKNAME": "nickname",
    "BDAY AGE": "bday_age",
    "BIRTHPLACE": "birthplace",
    "CIVIL STATUS": "civil_status",
    "NO OF YEARS MARRIED OR SEPARATED": "years_married_or_separated",
    "NATIONALITY": "nationality",
    "EDUCATION ATTAINMENT": "education",
    "EDUCATIONAL ATTAINMENT": "education",
}


SECTION_TOKENS = {
    "NAME OF SUBJECT": "subject",
    "NAME OF SPOUSE LIVE IN PARTNER": "spouse",
}


INCOME_FIELD_MAP = {
    "EMPLOYER NAME": "employer_name",
    "ADDRESS": "employer_address",
    "POSITION": "position",
    "TENURE LENGTH OF SERVICE": "tenure",
    "SALARY": "salary",
    "STATUS": "employment_status",
    "IF OFW SEAMAN ASK IF CURRENTLY ON BOARD ON VACATION": "ofw_status",
    "DEPARTURE DATE": "departure_date",
    "ARRIVAL DATE": "arrival_date",
    "BUSINESS NAME": "business_name",
    "YEARS OF OPERATION": "business_years_of_operation",
    "MONTHLY GROSS INCOME": "monthly_gross_income",
    "MONTHLY NET INCOME": "monthly_net_income",
    "MONTHLY REMITTANCE ALLOTMENT AMOUNT": "monthly_remittance_amount",
    "MONTHLY REMITTANCE AMOUNT": "monthly_remittance_amount",
    "BANK REMITTANCE CENTER": "remittance_center",
    "ADDRESS BRANCH": "remittance_branch",
    "SENDER S NAME": "sender_name",
    "RELATIONSHIP TO SENDER": "relationship_to_sender",
}


SECTION_HEADINGS = {
    "SOURCES OF INCOME OF SUBJECT": "subject_income",
    "SOURCES OF INCOME OF SPOUSE": "spouse_income",
    "VEHICLES": "vehicle",
    "INFORMANTS REMARKS": "informants",
}


INFORMANT_SLOT_MAP = {
    "BARANGAY INFORMANT S NAME": "barangay_informant",
    "NEIGHBOR S NAME": "neighbor",
    "NEIGHBOR NAME": "neighbor",
    "MAIN INFORMANT S NAME": "main_informant",
}


INFORMANT_FIELD_MAP = {
    "BARANGAY INFORMANT S NAME": "name",
    "POSITION": "position",
    "AGE": "age",
    "CONTACT NUMBER": "contact_number",
    "BARANGAY ADDRESS": "address",
    "IF SUBJECT COMAKER IS KNOWN OR UNKNOWN": "known",
    "IF SUBJECT COMAKER IS VOTER OR NOT VOTER": "voter",
    "IF SUBJECT COMAKER IS RESIDENT OR NOT RESIDENT": "resident",
    "IF SUBJECT COMAKER HAS BAD RECORD": "has_bad_record",
    "IF THE GIVEN ADDRESS IS THE SAME IN RECORD": "same_address",
    "RECORD IS BASED ON": "record_based_on",
    "SELFIE PICTURE": "selfie_picture",
    "PICTURE OF LOGBOOK": "logbook_picture",
    "PICTURE OF VOTER S RECORD": "voter_record_picture",
    "NEIGHBOR S NAME": "name",
    "NEIGHBOR NAME": "name",
    "RELATION TO SUBJECT COMAKER": "relationship",
    "YEARS IN THE AREA": "years_in_area",
    "YEARS IN THE AREA OF NEIGHBOR": "years_in_area",
    "ADDRESS": "address",
    "MAIN INFORMANT S NAME": "name",
    "RELATION TO SUBJECT": "relationship",
    "REMARKS": "remarks",
}


def normalize_label(value: str) -> str:
    cleaned = re.sub(r"[^A-Z0-9]+", " ", value.upper()).strip()
    return re.sub(r"\s+", " ", cleaned)


def compact_value(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()


def join_name(last_name: str, first_name: str, middle_name: str) -> str:
    last_name = compact_value(last_name)
    first_name = compact_value(first_name)
    middle_name = compact_value(middle_name)
    first_part = " ".join(part for part in [first_name, middle_name] if part)
    if last_name and first_part:
        return f"{last_name}, {first_part}"
    return compact_value(" ".join(part for part in [last_name, first_name, middle_name] if part))


def split_address(address: str) -> Dict[str, str]:
    raw = compact_value(address)
    result = {
        "address_line_1": raw,
        "address_line_2": "",
        "address_line_3": "",
        "barangay": "",
        "city": "",
        "region": "",
    }
    if not raw:
        return result

    city_match = re.search(r"\b([A-Z ]+CITY)\b", raw.upper())
    if city_match:
        result["city"] = city_match.group(1).title()

    barangay_match = re.search(r"\b(BRGY\.?\s+[A-Z0-9 ]+?)(?=\s+[A-Z ]+CITY|$)", raw.upper())
    if barangay_match:
        result["barangay"] = compact_value(barangay_match.group(1).replace("BRGY.", "BRGY"))

    street_part = raw
    if result["barangay"]:
        street_part = re.split(r"\bBRGY\.?\b", raw, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    elif result["city"]:
        street_part = raw[: raw.upper().find(result["city"].upper())].strip()

    result["address_line_2"] = street_part
    return result


def split_maybank_address(address: str) -> Dict[str, str]:
    raw = compact_value(address)
    result = {
        "unit_building": "",
        "street": "",
        "village_subdivision": "",
        "barangay": "",
        "city": "",
        "province": "",
        "region": "",
    }
    if not raw:
        return result

    parts = [compact_value(part) for part in raw.split(",") if compact_value(part)]
    first_part = parts[0] if parts else raw
    first_match = re.match(r"^([0-9A-Z\-\/]+)\s+(.+)$", first_part, flags=re.IGNORECASE)
    if first_match:
        result["unit_building"] = compact_value(first_match.group(1))
        result["street"] = compact_value(first_match.group(2))
    else:
        result["street"] = first_part

    leftovers: List[str] = []
    for part in parts[1:]:
        normalized = normalize_label(part)
        if "BRGY" in normalized or "BARANGAY" in normalized:
            result["barangay"] = part
        elif "CITY" in normalized or "MUNICIPALITY" in normalized:
            result["city"] = part
        elif not result["province"] and len(parts) >= 2 and part == parts[-1]:
            result["province"] = part
        else:
            leftovers.append(part)

    if leftovers:
        result["village_subdivision"] = leftovers[0]
        if len(leftovers) > 1 and not result["province"]:
            result["province"] = leftovers[-1]

    return result


def infer_dependents(dependents_text: str) -> Tuple[str, str]:
    value = compact_value(dependents_text)
    if not value or normalize_label(value) == "NONE":
        return "0", "None"

    entries = [item.strip() for item in re.split(r"[;,]", value) if item.strip()]
    ages = []
    for entry in entries:
        age_match = re.search(r"\b(\d{1,2})\b", entry)
        if age_match:
            ages.append(age_match.group(1))
    return str(len(entries)), ", ".join(ages)


def parse_dependent_rows(dependents_text: str) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for raw_line in dependents_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if normalize_label(line) == "NONE":
            continue
        line = re.sub(r"^\d+[\.\)]?\s*", "", line).strip()
        if not line:
            continue

        segments = [compact_value(segment) for segment in line.split("/") if compact_value(segment)]
        rows.append(
            {
                "name": segments[0] if len(segments) > 0 else line,
                "age": segments[1] if len(segments) > 1 else "",
                "school": segments[2] if len(segments) > 2 else "",
                "grade": segments[3] if len(segments) > 3 else "",
                "course": segments[4] if len(segments) > 4 else "",
            }
        )
    return rows[:3]


def infer_utility_bills(record: Dict[str, Any]) -> str:
    if record.get("utility_bills"):
        return record["utility_bills"]

    raw = record.get("raw_notes", "").lower()
    if "meralco bill" in raw:
        return "Electricity Bill (Meralco)"
    return "None"


def infer_outcome(record: Dict[str, Any]) -> str:
    if record.get("outcome"):
        return record["outcome"]

    remarks = record.get("remarks", "").lower()
    if "confirmed subject residing" in remarks or "confirmed given address" in remarks:
        return "VERIFIED"
    return "PARTIALLY VERIFIED"


def infer_adverse_finding(record: Dict[str, Any]) -> str:
    if record.get("adverse_finding"):
        return record["adverse_finding"]

    remarks = record.get("remarks", "").lower()
    if "unknown ang subject" in remarks or "not known" in remarks:
        return "Not known to neighborhood"
    return "No Adverse Findings"


def infer_living_condition(record: Dict[str, Any]) -> str:
    if record.get("living_condition"):
        return record["living_condition"]

    grade = compact_value(record.get("classification", "") or record.get("house_condition", "")).lower()
    if any(token in grade for token in ["excellent", "very good"]):
        return "Wealthy Living Standard"
    if any(token in grade for token in ["poor", "very bad"]):
        return "Poor Living Standard"
    return "Common Living Standard"


def infer_neighborhood_classification(record: Dict[str, Any]) -> str:
    if record.get("neighborhood_classification"):
        return record["neighborhood_classification"]

    neighborhood = compact_value(record.get("neighborhood", "")).lower()
    if "residential" in neighborhood:
        return "Residential"
    if "commercial" in neighborhood:
        return "Commercial"
    return "Others"


def infer_area_definition(record: Dict[str, Any]) -> str:
    if record.get("area_definition"):
        return record["area_definition"]

    neighborhood = compact_value(record.get("neighborhood", "")).lower()
    accessibility = compact_value(record.get("accessibility", "")).lower()
    if "residential" in neighborhood:
        if "mix" in accessibility:
            return "Mixed Average With Some Parts Poor Residential Area"
        return "Average Residential Area"
    return "Others"


def infer_ownership_verification(record: Dict[str, Any]) -> str:
    if record.get("ownership_verification"):
        return record["ownership_verification"]

    remarks = record.get("remarks", "").lower()
    if "confirmed given address" in remarks:
        return "Verified"
    return "Doubtful"


def infer_basis_of_ownership(record: Dict[str, Any]) -> str:
    if record.get("basis_of_ownership_verification"):
        return record["basis_of_ownership_verification"]

    notes = record.get("raw_notes", "").lower()
    if "brgy" in notes and "official" in notes:
        return "Barangay Officials"
    return "Neighbours"


def parse_named_informant(value: str) -> Dict[str, str]:
    lines = [line.strip() for line in value.splitlines() if line.strip()]
    if not lines:
        return {"name": "", "relationship": "", "contact": ""}

    primary = lines[0]
    secondary = " ".join(lines[1:]).strip()
    return {
        "name": primary,
        "relationship": secondary,
        "contact": "NP",
    }


def parse_informant_lines(value: str) -> List[Dict[str, str]]:
    records: List[Dict[str, str]] = []
    for raw_line in value.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"^\d+\.?\s*", "", line)
        segments = [segment.strip() for segment in line.split("/") if segment.strip()]
        name = segments[0] if segments else line
        relationship = segments[1] if len(segments) > 1 else ""
        location = " / ".join(segments[2:]) if len(segments) > 2 else ""
        records.append(
            {
                "name": name,
                "relationship": relationship,
                "address": location,
            }
        )
    return records[:4]


def clean_notes_text(notes: str) -> str:
    value = unescape(notes or "")
    value = value.replace("\xa0", " ")
    return value.replace("\r\n", "\n").replace("\r", "\n")


def strip_label_prefix(value: str) -> str:
    cleaned = normalize_label(value)
    return re.sub(r"^\d+\s*", "", cleaned).strip()


def set_record_value(record: Dict[str, Any], key: str, value: str, append: bool = False) -> None:
    if key not in record:
        return
    cleaned = value.strip()
    if not cleaned:
        return
    if append and record.get(key):
        record[key] = f"{record[key]}\n{cleaned}".strip()
        return
    record[key] = cleaned


def resolve_informant_record_key(slot: str, field_name: str) -> str:
    if slot == "barangay_informant":
        return {
            "name": "barangay_informant_name",
            "position": "barangay_informant_position",
            "age": "barangay_informant_age",
            "contact_number": "barangay_informant_contact_number",
            "address": "barangay_informant_address",
            "known": "barangay_subject_known",
            "voter": "barangay_subject_voter",
            "resident": "barangay_subject_resident",
            "has_bad_record": "barangay_has_bad_record",
            "same_address": "barangay_same_address",
            "record_based_on": "barangay_record_based_on",
            "selfie_picture": "barangay_selfie_picture",
            "logbook_picture": "barangay_logbook_picture",
            "voter_record_picture": "barangay_voter_record_picture",
            "remarks": "barangay_remarks",
        }.get(field_name, "")

    if slot.startswith("neighbor_"):
        return {
            "name": f"{slot}_name",
            "age": f"{slot}_age",
            "relationship": f"{slot}_relationship",
            "years_in_area": f"{slot}_years_in_area",
            "address": f"{slot}_address",
            "known": f"{slot}_known",
            "selfie_picture": f"{slot}_selfie_picture",
            "remarks": f"{slot}_remarks",
        }.get(field_name, "")

    if slot == "main_informant":
        return {
            "name": "main_informant_name",
            "address": "main_informant_address",
            "age": "main_informant_age",
            "relationship": "main_informant_relationship",
            "selfie_picture": "main_informant_selfie_picture",
            "remarks": "main_informant_remarks",
        }.get(field_name, "")

    return ""


def set_informant_value(record: Dict[str, Any], slot: str, field_name: str, value: str) -> None:
    target_key = resolve_informant_record_key(slot, field_name)
    if not target_key:
        return
    set_record_value(record, target_key, value, append=field_name == "remarks")


def build_structured_informants_summary(record: Dict[str, Any]) -> str:
    lines: List[str] = []

    if compact_value(record.get("barangay_informant_name", "")):
        barangay_parts = [
            record.get("barangay_informant_name", ""),
            record.get("barangay_informant_position", ""),
            record.get("barangay_subject_known", ""),
            record.get("barangay_informant_address", ""),
        ]
        lines.append(compact_value(" / ".join(part for part in barangay_parts if compact_value(part))))

    for index in range(1, 4):
        prefix = f"neighbor_{index}"
        if not compact_value(record.get(f"{prefix}_name", "")):
            continue
        neighbor_parts = [
            record.get(f"{prefix}_name", ""),
            record.get(f"{prefix}_relationship", ""),
            record.get(f"{prefix}_known", ""),
            record.get(f"{prefix}_address", ""),
        ]
        lines.append(compact_value(" / ".join(part for part in neighbor_parts if compact_value(part))))

    return "\n".join(line for line in lines if line).strip()


def build_structured_main_informant(record: Dict[str, Any]) -> str:
    if not compact_value(record.get("main_informant_name", "")):
        return ""

    detail_lines = [record.get("main_informant_name", "").strip()]
    extras = [
        record.get("main_informant_relationship", ""),
        record.get("main_informant_address", ""),
        record.get("main_informant_remarks", ""),
    ]
    extras = [compact_value(value) for value in extras if compact_value(value)]
    if extras:
        detail_lines.append(" | ".join(extras))
    return "\n".join(detail_lines).strip()


def build_maybank_informants(record: Dict[str, Any]) -> List[Dict[str, str]]:
    informants: List[Dict[str, str]] = []

    if compact_value(record.get("barangay_informant_name", "")):
        relationship = compact_value(
            " / ".join(
                value
                for value in [
                    "Barangay Informant",
                    record.get("barangay_informant_position", ""),
                    record.get("barangay_subject_known", ""),
                ]
                if compact_value(value)
            )
        )
        informants.append({"name": record.get("barangay_informant_name", ""), "relationship": relationship})

    for index in range(1, 4):
        prefix = f"neighbor_{index}"
        name = compact_value(record.get(f"{prefix}_name", ""))
        if not name:
            continue
        relationship = compact_value(
            " / ".join(
                value
                for value in [
                    record.get(f"{prefix}_relationship", ""),
                    record.get(f"{prefix}_known", ""),
                ]
                if compact_value(value)
            )
        )
        informants.append({"name": name, "relationship": relationship})

    if compact_value(record.get("main_informant_name", "")):
        relationship = compact_value(
            " / ".join(
                value
                for value in [
                    "Main Informant",
                    record.get("main_informant_relationship", ""),
                ]
                if compact_value(value)
            )
        )
        informants.append({"name": record.get("main_informant_name", ""), "relationship": relationship})

    if informants:
        return informants

    return parse_informant_lines(record.get("informants", ""))


def split_vehicle_unit_year_model(value: str) -> Dict[str, str]:
    raw = compact_value(value)
    if not raw:
        return {"make_model": "", "year_model": ""}

    match = re.search(r"\b(19|20)\d{2}\b", raw)
    if not match:
        return {"make_model": raw, "year_model": ""}

    year = match.group(0)
    make_model = compact_value(raw[: match.start()])
    trailing = compact_value(raw[match.end() :])
    if trailing:
        make_model = compact_value(f"{make_model} {trailing}")
    return {"make_model": make_model, "year_model": year}


def parse_ci_notes(notes: str) -> Dict[str, Any]:
    notes = clean_notes_text(notes)
    record = deepcopy(CI_FORM_DEFAULTS)
    record["raw_notes"] = notes.strip()

    current_person = "subject"
    current_income_person = ""
    current_income_mode = ""
    current_informant_slot = ""
    current_field_label = ""
    current_field_lines: List[str] = []
    neighbor_counter = 0

    def commit_field() -> None:
        nonlocal current_field_label, current_field_lines, current_informant_slot, neighbor_counter
        if not current_field_label:
            return

        label_base = strip_label_prefix(current_field_label)
        value = "\n".join(line for line in current_field_lines if line.strip()).strip()
        if re.fullmatch(r"[.\-_/ ]*", value or ""):
            value = ""

        slot_type = INFORMANT_SLOT_MAP.get(label_base)
        if label_base.startswith("MAIN INFORMANT"):
            slot_type = "main_informant"
        if slot_type == "barangay_informant":
            current_informant_slot = "barangay_informant"
        elif slot_type == "neighbor":
            neighbor_counter += 1
            current_informant_slot = f"neighbor_{min(neighbor_counter, 3)}"
        elif slot_type == "main_informant":
            current_informant_slot = "main_informant"

        if label_base.startswith("DEPENDENTS AGE SCHOOL"):
            set_record_value(record, "dependents", value, append=bool(record.get("dependents")))
        elif label_base == "INFORMANTS":
            set_record_value(record, "informants", value, append=bool(record.get("informants")))
        elif label_base.startswith("MAIN INFORMANT") and not current_informant_slot:
            set_record_value(record, "main_informant", value, append=bool(record.get("main_informant")))
        elif current_informant_slot and label_base in INFORMANT_FIELD_MAP:
            set_informant_value(record, current_informant_slot, INFORMANT_FIELD_MAP[label_base], value)
        elif label_base == "REMARKS":
            if current_informant_slot:
                set_informant_value(record, current_informant_slot, "remarks", value)
            else:
                set_record_value(record, "remarks", value, append=bool(record.get("remarks")))
        elif label_base in PERSON_FIELD_MAP:
            field_name = f"{current_person}_{PERSON_FIELD_MAP[label_base]}"
            set_record_value(record, field_name, value)
        elif label_base in {"EMPLOYMENT", "BUSINESS"}:
            target_person = current_income_person or current_person
            set_record_value(record, f"{target_person}_{label_base.lower()}", value)
        elif current_income_person:
            income_field = INCOME_FIELD_MAP.get(label_base)
            if income_field:
                set_record_value(record, f"{current_income_person}_{income_field}", value)
                if label_base == "EMPLOYER NAME":
                    set_record_value(record, f"{current_income_person}_employment", value)
                if label_base == "BUSINESS NAME":
                    set_record_value(record, f"{current_income_person}_business", value)
            elif label_base == "ADDRESS":
                if current_income_mode == "business":
                    set_record_value(record, f"{current_income_person}_business_address", value)
                else:
                    set_record_value(record, f"{current_income_person}_employer_address", value)
            else:
                target_key = DIRECT_FIELD_MAP.get(label_base)
                if target_key:
                    set_record_value(record, target_key, value)
        else:
            target_key = DIRECT_FIELD_MAP.get(label_base)
            if target_key:
                set_record_value(record, target_key, value)

        if label_base.startswith("MAIN INFORMANT") and current_informant_slot == "main_informant":
            set_record_value(record, "main_informant", value, append=bool(record.get("main_informant")))

        current_field_label = ""
        current_field_lines = []

    for raw_line in notes.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        if re.fullmatch(r"[-=*_\s]+", line):
            commit_field()
            continue

        normalized = strip_label_prefix(line.rstrip(":"))
        if normalized in {"REMINDER", "IF RENTAL", "FOR BDO ACCOUNT", "FOR ALL ACCOUNT", "FOR CBS ACCOUNT", "PHYSICAL ATTRIBUTES", "PHYSICAL ATTRIBUTES FOR PNB"}:
            commit_field()
            continue

        if normalized in SECTION_TOKENS:
            commit_field()
            current_person = SECTION_TOKENS[normalized]
            current_income_person = ""
            current_income_mode = ""
            current_informant_slot = ""
            continue

        section_name = SECTION_HEADINGS.get(normalized)
        if section_name:
            commit_field()
            current_informant_slot = ""
            if section_name == "subject_income":
                current_income_person = "subject"
                current_income_mode = ""
            elif section_name == "spouse_income":
                current_income_person = "spouse"
                current_income_mode = ""
            else:
                current_income_person = ""
                current_income_mode = ""
            continue

        if normalized == "SOURCES OF INCOME":
            commit_field()
            current_income_person = current_person
            current_income_mode = ""
            continue

        if ":" not in line and normalized in {"IF EMPLOYMENT", "EMPLOYMENT"}:
            commit_field()
            if current_income_person:
                current_income_mode = "employment"
            continue

        if ":" not in line and normalized in {"IF BUSINESS", "BUSINESS"}:
            commit_field()
            if current_income_person:
                current_income_mode = "business"
            continue

        if ":" not in line and normalized in {"IF REMITTANCE", "IF REMITTANCE ALLOTMENT"}:
            commit_field()
            if current_income_person:
                current_income_mode = "remittance"
            continue

        if normalized.startswith("DEPENDENTS AGE SCHOOL"):
            commit_field()
            current_field_label = "DEPENDENTS AGE SCHOOL"
            current_field_lines = []
            continue

        if ":" in line:
            commit_field()
            label, value = [part.strip() for part in line.split(":", 1)]
            current_field_label = label
            current_field_lines = [value] if value else []
            continue

        if current_field_label:
            current_field_lines.append(line)
            continue

    commit_field()

    if not record["complete_address"]:
        record["complete_address"] = record.get("verified_address") or record.get("given_address", "")
    if not record["present_address"]:
        record["present_address"] = record.get("verified_address") or record.get("given_address", "")
    if not record["vehicles"] and record.get("vehicle_unit_year_model"):
        record["vehicles"] = record["vehicle_unit_year_model"]
    if not record["remarks"] and record.get("ci_remarks_observation"):
        record["remarks"] = record["ci_remarks_observation"]
    if not record["informants"]:
        record["informants"] = build_structured_informants_summary(record)
    if not record["main_informant"]:
        record["main_informant"] = build_structured_main_informant(record)

    if not record["field_investigator"]:
        record["field_investigator"] = "CI User"
    if not record["submitted_date"]:
        record["submitted_date"] = datetime.now().strftime("%Y-%m-%d")
    if not record["submitted_time"]:
        record["submitted_time"] = datetime.now().strftime("%I:%M %p")
    if not record["date_time_inspected"]:
        record["date_time_inspected"] = f"{record['submitted_date']} {record.get('time_of_visit', '').strip()}".strip()

    return record


def build_export_payload(record: Dict[str, Any]) -> Dict[str, Any]:
    payload = deepcopy(record)
    payload["applicant_name"] = join_name(
        payload.get("subject_last_name", ""),
        payload.get("subject_first_name", ""),
        payload.get("subject_middle_name", ""),
    )
    payload["spouse_name"] = join_name(
        payload.get("spouse_last_name", ""),
        payload.get("spouse_first_name", ""),
        payload.get("spouse_middle_name", ""),
    )
    dependents_count, dependents_ages = infer_dependents(payload.get("dependents", ""))
    payload["dependents_count"] = dependents_count
    payload["dependents_ages"] = dependents_ages
    payload["dependent_rows"] = parse_dependent_rows(payload.get("dependents", ""))
    address_parts = split_address(payload.get("complete_address", ""))
    payload.update(address_parts)
    payload["maybank_address_parts"] = split_maybank_address(
        payload.get("verified_address", "") or payload.get("complete_address", "") or payload.get("given_address", "")
    )
    payload["utility_bills"] = infer_utility_bills(payload)
    payload["living_condition"] = infer_living_condition(payload)
    payload["neighborhood_classification"] = infer_neighborhood_classification(payload)
    payload["area_definition"] = infer_area_definition(payload)
    payload["ownership_verification"] = infer_ownership_verification(payload)
    payload["basis_of_ownership_verification"] = infer_basis_of_ownership(payload)
    payload["adverse_finding"] = infer_adverse_finding(payload)
    payload["outcome"] = infer_outcome(payload)
    payload["main_informant_details"] = parse_named_informant(payload.get("main_informant", ""))
    payload["informant_list"] = parse_informant_lines(payload.get("informants", ""))
    payload["maybank_informants"] = build_maybank_informants(payload)
    payload["vehicle_details"] = split_vehicle_unit_year_model(payload.get("vehicle_unit_year_model", "") or payload.get("vehicles", ""))
    payload["address_type"] = payload.get("address_type") or "Residential Address"
    payload["type_of_residence_export"] = "House and lot" if payload.get("type_of_residence") else ""
    payload["subject_income"] = " / ".join(
        [value for value in [payload.get("subject_employment", ""), payload.get("subject_business", "")] if compact_value(value)]
    )
    payload["spouse_income"] = " / ".join(
        [value for value in [payload.get("spouse_employment", ""), payload.get("spouse_business", "")] if compact_value(value)]
    )
    payload["request_reference"] = (
        payload.get("reference_number")
        or payload.get("los_request_number")
        or payload.get("bpi_request_number")
        or payload.get("account_name")
    )
    payload["date_time_submitted"] = f"{payload.get('submitted_date', '')} {payload.get('submitted_time', '')}".strip()
    payload["residence_description"] = compact_value(
        " | ".join(
            value
            for value in [
                payload.get("type_of_residence", ""),
                payload.get("no_of_storey", ""),
                payload.get("classification", ""),
                payload.get("house_condition", ""),
                payload.get("made_of", ""),
            ]
            if compact_value(value)
        )
    )
    payload["general_appearance"] = compact_value(
        " | ".join(
            value
            for value in [
                payload.get("appearance", ""),
                payload.get("house_color", ""),
                payload.get("gate_color", ""),
                payload.get("fence_color", ""),
                payload.get("exterior", ""),
            ]
            if compact_value(value)
        )
    )
    return payload


def get_available_templates() -> Dict[str, Dict[str, Any]]:
    return TEMPLATE_DEFINITIONS


def get_template_workbook(template_key: str = DEFAULT_TEMPLATE_KEY, template_bytes: bytes | None = None):
    if template_bytes:
        return load_workbook(BytesIO(template_bytes))

    template_path = TEMPLATE_DEFINITIONS[template_key]["path"]
    with open(template_path, "rb") as handle:
        return load_workbook(BytesIO(handle.read()))


def set_value(sheet, cell_ref: str, value: Any) -> None:
    target_ref = cell_ref
    for merged_range in sheet.merged_cells.ranges:
        if cell_ref in merged_range:
            target_ref = merged_range.start_cell.coordinate
            break
    sheet[target_ref] = "" if value is None else value


def fill_bpi_template(workbook, payload: Dict[str, Any]) -> None:
    sheet = workbook[TEMPLATE_DEFINITIONS["bpi"]["sheet_name"]]
    cell_map = {
        "P4": payload.get("bpi_request_number", ""),
        "P5": payload.get("los_request_number", ""),
        "M8": payload.get("applicant_name", ""),
        "M9": payload.get("subject_civil_status", ""),
        "M10": payload.get("spouse_name", ""),
        "M11": payload.get("dependents_count", ""),
        "AU11": payload.get("dependents_ages", ""),
        "AA13": payload.get("address_type", ""),
        "AA14": payload.get("address_line_1", ""),
        "AA15": payload.get("address_line_2", ""),
        "AA16": payload.get("address_line_3", ""),
        "AA17": payload.get("barangay", ""),
        "AA18": payload.get("city", ""),
        "AA19": payload.get("province", ""),
        "AA20": payload.get("present_address", ""),
        "AA21": payload.get("previous_address", ""),
        "AA22": payload.get("utility_bills", ""),
        "AA23": payload.get("coordinates", ""),
        "U25": payload.get("length_of_stay", ""),
        "U26": payload.get("ownership", ""),
        "U27": payload.get("ownership_verification", ""),
        "U28": payload.get("basis_of_ownership_verification", ""),
        "U29": payload.get("type_of_residence_export", ""),
        "U30": payload.get("living_condition", ""),
        "U32": payload.get("neighborhood_classification", ""),
        "U33": payload.get("area_definition", ""),
        "BA25": payload.get("no_of_storey", ""),
        "BA26": payload.get("parking", ""),
        "AW35": payload.get("adverse_finding", ""),
        "AW36": payload.get("time_of_visit", ""),
        "B39": payload.get("remarks", ""),
        "T43": payload["main_informant_details"].get("name", ""),
        "T44": payload["main_informant_details"].get("relationship", ""),
        "T45": payload["main_informant_details"].get("contact", ""),
        "T52": payload.get("outcome", ""),
        "T56": payload.get("field_investigator", ""),
        "U58": payload.get("submitted_date", ""),
        "AC58": payload.get("submitted_time", ""),
    }

    for cell_ref, value in cell_map.items():
        set_value(sheet, cell_ref, value)

    informant_rows = [47, 48, 49, 50]
    for row_number, informant in zip(informant_rows, payload["informant_list"]):
        set_value(sheet, f"D{row_number}", informant.get("name", ""))
        set_value(sheet, f"V{row_number}", informant.get("relationship", ""))
        set_value(sheet, f"AI{row_number}", informant.get("address", ""))


def fill_maybank_template(workbook, payload: Dict[str, Any]) -> None:
    sheet = workbook[TEMPLATE_DEFINITIONS["maybank"]["sheet_name"]]
    address_parts = payload.get("maybank_address_parts", {})
    vehicle_details = payload.get("vehicle_details", {})
    cell_map = {
        "J10": payload.get("request_reference", ""),
        "AB10": payload.get("date_time_inspected", ""),
        "J12": payload.get("requesting_officer", ""),
        "AB12": payload.get("date_time_submitted", ""),
        "F16": payload.get("applicant_name", ""),
        "Q16": payload.get("subject_bday_age", ""),
        "V16": payload.get("subject_education", ""),
        "AG16": payload.get("subject_income", ""),
        "F18": payload.get("subject_civil_status", ""),
        "Q18": payload.get("subject_gender", ""),
        "V18": payload.get("subject_nationality", ""),
        "AG18": payload.get("dependents_count", ""),
        "F20": payload.get("spouse_name", ""),
        "Q20": payload.get("spouse_bday_age", ""),
        "V20": payload.get("spouse_education", ""),
        "AG20": payload.get("spouse_income", ""),
        "F22": payload.get("spouse_civil_status", ""),
        "Q22": payload.get("spouse_gender", ""),
        "V22": payload.get("spouse_nationality", ""),
        "B25": payload["dependent_rows"][0]["name"] if len(payload["dependent_rows"]) > 0 else "",
        "J25": payload["dependent_rows"][0]["age"] if len(payload["dependent_rows"]) > 0 else "",
        "O25": payload["dependent_rows"][0]["school"] if len(payload["dependent_rows"]) > 0 else "",
        "Y25": payload["dependent_rows"][0]["grade"] if len(payload["dependent_rows"]) > 0 else "",
        "AE25": payload["dependent_rows"][0]["course"] if len(payload["dependent_rows"]) > 0 else "",
        "B26": payload["dependent_rows"][1]["name"] if len(payload["dependent_rows"]) > 1 else "",
        "J26": payload["dependent_rows"][1]["age"] if len(payload["dependent_rows"]) > 1 else "",
        "O26": payload["dependent_rows"][1]["school"] if len(payload["dependent_rows"]) > 1 else "",
        "Y26": payload["dependent_rows"][1]["grade"] if len(payload["dependent_rows"]) > 1 else "",
        "AE26": payload["dependent_rows"][1]["course"] if len(payload["dependent_rows"]) > 1 else "",
        "B29": address_parts.get("unit_building", ""),
        "H29": address_parts.get("street", ""),
        "N29": address_parts.get("village_subdivision", ""),
        "T29": address_parts.get("barangay", ""),
        "Y29": address_parts.get("city", ""),
        "AE29": address_parts.get("province", ""),
        "AJ29": address_parts.get("region", ""),
        "L32": payload.get("present_address", "") or payload.get("previous_address", ""),
        "B35": payload.get("ownership", ""),
        "AD36": payload.get("landlord_name", ""),
        "AD37": payload.get("rental_fee", "") or payload.get("vehicle_monthly_amortization", ""),
        "AD38": payload.get("outcome", ""),
        "E44": payload.get("length_of_stay", ""),
        "B47": payload.get("residence_description", ""),
        "W47": payload.get("contact_number", ""),
        "AF47": payload.get("subject_income", "") or payload.get("ownership", ""),
        "P49": payload.get("lot_area", ""),
        "P50": payload.get("floor_area", ""),
        "B56": payload.get("general_appearance", ""),
        "B59": payload.get("utility_bills", "") or payload.get("basis_of_ownership_verification", ""),
        "K59": payload.get("adverse_finding", "") or payload.get("remarks", ""),
        "W59": payload.get("neighborhood_classification", ""),
        "B65": payload.get("neighborhood", ""),
        "P65": payload.get("accessibility", ""),
        "W66": payload.get("landmark", ""),
        "AG66": payload.get("bdo_distance", ""),
        "W67": payload.get("corner", ""),
        "W68": payload.get("time_of_visit", ""),
        "W69": payload.get("security_guard_hoa", ""),
        "B74": vehicle_details.get("make_model", ""),
        "L74": vehicle_details.get("year_model", ""),
        "Q74": payload.get("vehicle_owned_or_mortgage", ""),
        "V74": payload.get("ownership", "") if compact_value(payload.get("vehicle_owned_or_mortgage", "")).lower().startswith("mort") else "",
        "AC74": payload.get("vehicle_monthly_amortization", ""),
        "B79": payload.get("ci_remarks_observation", "") or payload.get("remarks", ""),
        "B103": payload.get("field_investigator", ""),
        "T103": payload.get("requesting_officer", ""),
    }

    for cell_ref, value in cell_map.items():
        set_value(sheet, cell_ref, value)

    informant_rows = [91, 92, 93, 94, 95, 96, 97, 98]
    all_informants = payload["maybank_informants"][:8]
    for row_number, informant in zip(informant_rows, all_informants):
        set_value(sheet, f"B{row_number}", informant.get("name", ""))
        set_value(sheet, f"T{row_number}", informant.get("relationship", ""))


def export_ci_to_template(
    record: Dict[str, Any],
    template_key: str = DEFAULT_TEMPLATE_KEY,
    template_bytes: bytes | None = None,
) -> Tuple[bytes, str]:
    payload = build_export_payload(record)
    workbook = get_template_workbook(template_key=template_key, template_bytes=template_bytes)

    if template_key == "maybank":
        fill_maybank_template(workbook, payload)
    else:
        fill_bpi_template(workbook, payload)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    applicant_slug = re.sub(r"[^A-Za-z0-9]+", "_", payload.get("subject_last_name", "record")).strip("_") or "record"
    filename = f"{template_key}_{applicant_slug}_pdrn_output.xlsx"
    return output.getvalue(), filename


def save_export_copy(file_bytes: bytes, filename: str) -> str:
    export_dir = Path.cwd() / "exports"
    export_dir.mkdir(exist_ok=True)
    output_path = export_dir / filename
    output_path.write_bytes(file_bytes)
    return str(output_path)
